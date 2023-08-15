import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox
from calendar import monthrange
from datetime import datetime

ano_atual = datetime.now().year

def adicionar_dados():
    data = f"{var_ano.get()}-{var_mes.get()}-{var_dia.get()}"
    descricao = entry_descricao.get()
    valor = float(entry_valor.get())
    cartao = entry_cartao.get()
    parcelas = int(entry_parcelas.get())
    tipo = entry_tipo.get()

    nova_linha = [data, descricao, valor, cartao, parcelas, tipo]
    financas_page.append(nova_linha)
    book.save('Pasta.xlsx')

    messagebox.showinfo("Sucesso", "Dados adicionados com sucesso!")

    # Limpa os campos de entrada após adicionar os dados
    entry_descricao.delete(0, tk.END)
    entry_valor.delete(0, tk.END)
    entry_cartao.delete(0, tk.END)
    entry_parcelas.delete(0, tk.END)
    entry_tipo.delete(0, tk.END)
    
    atualizar_soma()
    atualizar_treeview()

    
def calcular_soma():
    soma = sum(float(row[2]) for row in financas_page.iter_rows(min_row=2, values_only=True))
    return soma

def atualizar_soma():
    soma = calcular_soma()
    label_soma.config(text=f"Soma dos valores: R${soma: .2f}")
    

def atualizar_treeview():
    for item in tree.get_children():
        tree.delete(item)

    for row in financas_page.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)
    atualizar_soma() #Atualiza a soma dos valores

def atualizar_dias(*args):
    ano = var_ano.get()
    mes = var_mes.get()

    ultimo_dia = monthrange(ano, mes)[1]
    dias_disponiveis = list(range(1, ultimo_dia + 1))

    var_dia.set(dias_disponiveis[0])  # Define o dia para o primeiro disponível
    menu_dia['values'] = dias_disponiveis

book = openpyxl.load_workbook('Pasta.xlsx')
financas_page = book['Planilha1']

root = tk.Tk()
root.title("Registro de Finanças")


# Labels
label_data = tk.Label(root, text="Data:")
label_descricao = tk.Label(root, text="Descrição:")
label_valor = tk.Label(root, text="Valor:")
label_cartao = tk.Label(root, text="Cartão:")
label_parcelas = tk.Label(root, text="Parcelas:")
label_tipo = tk.Label(root, text="Tipo:")

label_data.grid(row=0, column=0)
label_descricao.grid(row=1, column=0)
label_valor.grid(row=2, column=0)
label_cartao.grid(row=3, column=0)
label_parcelas.grid(row=4, column=0)
label_tipo.grid(row=5, column=0)

# Label para mostrar a soma dos valores
label_soma = tk.Label(root, text="Soma dos Valores: R$ 0.00")
label_soma.grid(row=9, column=0, columnspan=2)



# Entry fields
entry_descricao = tk.Entry(root)
entry_valor = tk.Entry(root)
entry_cartao = tk.Entry(root)
entry_parcelas = tk.Entry(root)
entry_tipo = tk.Entry(root)

entry_descricao.grid(row=1, column=1)
entry_valor.grid(row=2, column=1)
entry_cartao.grid(row=3, column=1)
entry_parcelas.grid(row=4, column=1)
entry_tipo.grid(row=5, column=1)


# Labels
label_data = tk.Label(root, text="Data:")
label_descricao = tk.Label(root, text="Descrição:")
label_valor = tk.Label(root, text="Valor:")
label_cartao = tk.Label(root, text="Cartão:")
label_parcelas = tk.Label(root, text="Parcelas:")
label_tipo = tk.Label(root, text="Tipo:")

label_data.grid(row=0, column=0)
label_descricao.grid(row=1, column=0)
label_valor.grid(row=2, column=0)
label_cartao.grid(row=3, column=0)
label_parcelas.grid(row=4, column=0)
label_tipo.grid(row=5, column=0)

# Entry fields
entry_descricao = tk.Entry(root)
entry_valor = tk.Entry(root)
entry_cartao = tk.Entry(root)
entry_parcelas = tk.Entry(root)
entry_tipo = tk.Entry(root)

entry_descricao.grid(row=1, column=1)
entry_valor.grid(row=2, column=1)
entry_cartao.grid(row=3, column=1)
entry_parcelas.grid(row=4, column=1)
entry_tipo.grid(row=5, column=1)


# Variáveis de controle para os menus suspensos
var_dia = tk.StringVar(value="01")  # Valor padrão para o dia
var_mes = tk.StringVar(value="01")  # Valor padrão para o mês
var_ano = tk.StringVar(value="2023")  # Valor padrão para o ano


# Frame para os menus suspensos de data
frame_data = tk.Frame(root)
frame_data.grid(row=0, column=1, padx=5, pady=5)

# Menus suspensos (combobox)
menu_dia = ttk.Combobox(frame_data, textvariable=var_dia, values=[str(i).zfill(2) for i in range(1, 32)], state="readonly", width=3)
menu_dia.pack(side="left")
menu_mes = ttk.Combobox(frame_data, textvariable=var_mes, values=[str(i).zfill(2) for i in range(1, 13)], state="readonly", width=3)
menu_mes.pack(side="left")
menu_ano = ttk.Combobox(frame_data, textvariable=var_ano, values=[str(i) for i in range(ano_atual - 2, ano_atual + 3)], state="readonly", width=6)
menu_ano.pack(side="left")



# Treeview para mostrar os dados em formato de planilha
tree = ttk.Treeview(root, columns=("Data", "Descrição", "Valor", "Cartão", "Parcelas", "Tipo"), show="headings")
tree.heading("Data", text="Data")
tree.heading("Descrição", text="Descrição")
tree.heading("Valor", text="Valor")
tree.heading("Cartão", text="Cartão")
tree.heading("Parcelas", text="Parcelas")
tree.heading("Tipo", text="Tipo")
tree.grid(row=7, columnspan=2)

# Atualiza o Treeview com os dados da planilha
atualizar_treeview()


# Button
btn_adicionar = tk.Button(root, text="Adicionar Dados", command=adicionar_dados)
btn_adicionar.grid(row=8, columnspan=2)



root.mainloop()